from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///learning_website.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    clickstream_events = db.relationship('ClickstreamEvent', backref='user', lazy=True)
    quiz_attempts = db.relationship('QuizAttempt', backref='user', lazy=True)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    lessons = db.relationship('Lesson', backref='course', lazy=True)

class Lesson(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content_type = db.Column(db.String(50), nullable=False)  # 'text', 'video', 'quiz'
    content = db.Column(db.Text)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    order = db.Column(db.Integer, default=0)
    
    # Relationships
    quiz_questions = db.relationship('QuizQuestion', backref='lesson', lazy=True)

class QuizQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question = db.Column(db.Text, nullable=False)
    options = db.Column(db.Text)  # JSON string of options
    correct_answer = db.Column(db.Integer, nullable=False)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)

class QuizAttempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    score = db.Column(db.Float, nullable=False)
    answers = db.Column(db.Text)  # JSON string of user answers
    completed_at = db.Column(db.DateTime, default=datetime.utcnow)

class ClickstreamEvent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Nullable for anonymous users
    session_id = db.Column(db.String(100), nullable=False)
    event_type = db.Column(db.String(50), nullable=False)  # 'page_view', 'click', 'video_action', 'quiz_action'
    element_id = db.Column(db.String(100), nullable=True)
    element_type = db.Column(db.String(50), nullable=True)  # 'button', 'link', 'video', 'quiz_question'
    page_url = db.Column(db.String(500), nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    additional_data = db.Column(db.Text)  # JSON string for extra data
    ip_address = db.Column(db.String(45), nullable=True)  # IPv4/IPv6 support
    
    # For video tracking
    video_id = db.Column(db.String(100), nullable=True)
    video_action = db.Column(db.String(50), nullable=True)  # 'play', 'pause', 'seek', 'complete'
    video_time = db.Column(db.Float, nullable=True)
    
    # For quiz tracking
    quiz_id = db.Column(db.String(100), nullable=True)
    question_id = db.Column(db.String(100), nullable=True)
    answer_selected = db.Column(db.String(100), nullable=True)

class AdminUser(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    # Try to find admin user first, then regular user
    admin_user = AdminUser.query.get(int(user_id))
    if admin_user:
        return admin_user
    return User.query.get(int(user_id))

# Routes
@app.route('/')
def index():
    track_event('page_view', 'homepage', 'page')
    courses = Course.query.all()
    return render_template('index.html', courses=courses)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists')
            return redirect(url_for('register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email already registered')
            return redirect(url_for('register'))
        
        user = User(username=username, email=email, password_hash=generate_password_hash(password))
        db.session.add(user)
        db.session.commit()
        
        flash('Registration successful! Please login.')
        return redirect(url_for('login'))
    
    track_event('page_view', 'register_page', 'page')
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            track_event('login', 'login_button', 'button', user_id=user.id)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password')
    
    track_event('page_view', 'login_page', 'page')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    track_event('logout', 'logout_button', 'button')
    logout_user()
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    track_event('page_view', 'dashboard', 'page')
    courses = Course.query.all()
    return render_template('dashboard.html', courses=courses)

@app.route('/course/<int:course_id>')
@login_required
def course_detail(course_id):
    track_event('page_view', f'course_{course_id}', 'page')
    course = Course.query.get_or_404(course_id)
    lessons = Lesson.query.filter_by(course_id=course_id).order_by(Lesson.order).all()
    return render_template('course_detail.html', course=course, lessons=lessons)

@app.route('/lesson/<int:lesson_id>')
@login_required
def lesson_detail(lesson_id):
    track_event('page_view', f'lesson_{lesson_id}', 'page')
    lesson = Lesson.query.get_or_404(lesson_id)
    return render_template('lesson_detail.html', lesson=lesson)

@app.route('/api/track_event', methods=['POST'])
def api_track_event():
    """API endpoint for tracking events from frontend"""
    data = request.get_json()
    
    event_type = data.get('event_type')
    element_id = data.get('element_id')
    element_type = data.get('element_type')
    page_url = data.get('page_url')
    additional_data = data.get('additional_data', {})
    
    # Extract video-specific data
    video_id = additional_data.get('video_id')
    video_action = additional_data.get('video_action')
    video_time = additional_data.get('video_time')
    
    # Extract quiz-specific data
    quiz_id = additional_data.get('quiz_id')
    question_id = additional_data.get('question_id')
    answer_selected = additional_data.get('answer_selected')
    
    event = ClickstreamEvent(
        user_id=current_user.id if current_user.is_authenticated else None,
        session_id=session.get('session_id', 'anonymous'),
        event_type=event_type,
        element_id=element_id,
        element_type=element_type,
        page_url=page_url,
        additional_data=json.dumps(additional_data),
        ip_address=get_client_ip(),
        video_id=video_id,
        video_action=video_action,
        video_time=video_time,
        quiz_id=quiz_id,
        question_id=question_id,
        answer_selected=answer_selected
    )
    
    db.session.add(event)
    db.session.commit()
    
    return jsonify({'status': 'success'})

@app.route('/api/quiz-questions/<int:lesson_id>')
@login_required
def get_quiz_questions(lesson_id):
    """Get quiz questions for a specific lesson"""
    questions = QuizQuestion.query.filter_by(lesson_id=lesson_id).all()
    
    quiz_data = []
    for question in questions:
        quiz_data.append({
            'id': question.id,
            'question': question.question,
            'options': json.loads(question.options),
            'correct_answer': question.correct_answer
        })
    
    return jsonify({'questions': quiz_data})

@app.route('/api/submit_quiz', methods=['POST'])
@login_required
def submit_quiz():
    """Submit quiz answers and track results"""
    data = request.get_json()
    lesson_id = data.get('lesson_id')
    answers = data.get('answers')
    
    lesson = Lesson.query.get_or_404(lesson_id)
    questions = QuizQuestion.query.filter_by(lesson_id=lesson_id).all()
    
    correct_count = 0
    total_questions = len(questions)
    
    for i, question in enumerate(questions):
        user_answer = answers.get(str(i))
        if user_answer == question.correct_answer:
            correct_count += 1
        
        # Track each question attempt
        track_event('quiz_action', f'question_{question.id}', 'quiz_question', 
                   additional_data={
                       'quiz_id': lesson_id,
                       'question_id': question.id,
                       'answer_selected': user_answer,
                       'correct_answer': question.correct_answer
                   })
    
    score = (correct_count / total_questions) * 100
    
    # Save quiz attempt
    quiz_attempt = QuizAttempt(
        user_id=current_user.id,
        lesson_id=lesson_id,
        score=score,
        answers=json.dumps(answers)
    )
    db.session.add(quiz_attempt)
    db.session.commit()
    
    # Track quiz completion
    track_event('quiz_action', f'quiz_{lesson_id}_complete', 'quiz', 
               additional_data={'score': score, 'total_questions': total_questions})
    
    return jsonify({'score': score, 'correct_count': correct_count, 'total_questions': total_questions})

# Admin routes
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        admin_user = AdminUser.query.filter_by(username=username).first()
        if admin_user and check_password_hash(admin_user.password_hash, password):
            login_user(admin_user)
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    # Get summary statistics
    total_users = User.query.count()
    total_events = ClickstreamEvent.query.count()
    total_meaningful_events = ClickstreamEvent.query.filter(
        ~ClickstreamEvent.event_type.in_(['mouse_movement', 'visibility_change', 'time_on_page', 'scroll'])
    ).count()
    total_courses = Course.query.count()
    total_lessons = Lesson.query.count()
    
    # Get recent meaningful events (filter out excessive tracking and scroll events)
    meaningful_events = ClickstreamEvent.query.filter(
        ~ClickstreamEvent.event_type.in_(['mouse_movement', 'visibility_change', 'time_on_page', 'scroll'])
    ).order_by(ClickstreamEvent.timestamp.desc()).limit(20).all()
    
        # Get click analytics by category
    click_analytics = db.session.query(
        ClickstreamEvent.event_type,
        ClickstreamEvent.element_type,
        db.func.count(ClickstreamEvent.id).label('count')
    ).filter(
        ClickstreamEvent.event_type == 'click'
    ).group_by(
        ClickstreamEvent.event_type,
        ClickstreamEvent.element_type
    ).order_by(db.func.count(ClickstreamEvent.id).desc()).all()
    
    return render_template('admin_dashboard.html', 
                                                    total_users=total_users,
                           total_events=total_events,
                           total_meaningful_events=total_meaningful_events,
                           total_courses=total_courses,
                           total_lessons=total_lessons,
                           recent_events=meaningful_events,
                           click_analytics=click_analytics)

@app.route('/admin/export-excel')
@login_required
def export_excel():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    # Get meaningful clickstream events (filter out excessive tracking and scroll events)
    events = db.session.query(
        ClickstreamEvent, User.username, User.email
    ).outerjoin(User, ClickstreamEvent.user_id == User.id).filter(
        ~ClickstreamEvent.event_type.in_(['mouse_movement', 'visibility_change', 'time_on_page', 'scroll'])
    ).order_by(ClickstreamEvent.timestamp.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Analytics"
    
    # Define headers
    headers = ['Time', 'Event Context', 'Component', 'Event Name', 'Description', 'Origin', 'IP Address']
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row, (event, username, email) in enumerate(events, 2):
        # Time - More precise format
        ws.cell(row=row, column=1, value=event.timestamp.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
        
        # Event Context (page URL)
        ws.cell(row=row, column=2, value=event.page_url or 'N/A')
        
        # Component (element type)
        ws.cell(row=row, column=3, value=event.element_type or 'N/A')
        
        # Event Name (event type)
        ws.cell(row=row, column=4, value=event.event_type)
        
        # Description (additional data or element details)
        description = 'N/A'
        if event.additional_data:
            try:
                data = json.loads(event.additional_data)
                if isinstance(data, dict):
                    description = ', '.join([f"{k}: {v}" for k, v in data.items()])
                else:
                    description = str(data)
            except:
                description = event.additional_data
        elif event.element_id:
            description = f"Element: {event.element_id}"
        ws.cell(row=row, column=5, value=description)
        
        # Origin (user info)
        origin = 'Anonymous'
        if username:
            origin = f"User: {username} ({email})"
        ws.cell(row=row, column=6, value=origin)
        
        # IP Address
        ws.cell(row=row, column=7, value=event.ip_address or 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename=user_analytics.xlsx'
    }

@app.route('/admin/user-activity')
@login_required
def admin_user_activity():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    # Get all meaningful events with user information
    events = db.session.query(
        ClickstreamEvent, User.username, User.email
    ).outerjoin(User, ClickstreamEvent.user_id == User.id).filter(
        ~ClickstreamEvent.event_type.in_(['mouse_movement', 'visibility_change', 'time_on_page', 'scroll'])
    ).order_by(ClickstreamEvent.timestamp.desc()).all()
    
    # Process events to create detailed descriptions
    processed_events = []
    for event, username, email in events:
        # Create event context (page URL)
        event_context = event.page_url or 'N/A'
        if event.page_url:
            if '/course/' in event.page_url:
                course_id = event.page_url.split('/course/')[-1].split('/')[0]
                event_context = f"Course: {course_id}"
            elif '/lesson/' in event.page_url:
                lesson_id = event.page_url.split('/lesson/')[-1]
                event_context = f"Lesson: {lesson_id}"
            elif '/dashboard' in event.page_url:
                event_context = "Dashboard"
            elif '/admin/' in event.page_url:
                event_context = "Admin Panel"
            elif '/' == event.page_url or '/index' in event.page_url:
                event_context = "Homepage"
        
        # Create component type
        component = event.element_type or 'System'
        if event.element_type == 'button':
            component = 'Button'
        elif event.element_type == 'link':
            component = 'Navigation'
        elif event.element_type == 'form':
            component = 'Form'
        elif event.element_type == 'quiz_question':
            component = 'Quiz'
        elif event.element_type == 'video':
            component = 'Video'
        
        # Create event name
        event_name = event.event_type.replace('_', ' ').title()
        if event.event_type == 'page_view':
            event_name = 'Page Viewed'
        elif event.event_type == 'click':
            event_name = 'Element Clicked'
        elif event.event_type == 'login':
            event_name = 'User Login'
        elif event.event_type == 'logout':
            event_name = 'User Logout'
        elif event.event_type == 'quiz_action':
            event_name = 'Quiz Action'
        elif event.event_type == 'video_action':
            event_name = 'Video Action'
        
        # Create detailed description
        description = f"The user"
        if username:
            description += f" '{username}'"
        else:
            description += " (Anonymous)"
        
        if event.event_type == 'page_view':
            description += f" viewed the page '{event_context}'"
        elif event.event_type == 'click':
            description += f" clicked on a {component.lower()} element"
            if event.element_id:
                description += f" with id '{event.element_id}'"
        elif event.event_type == 'login':
            description += f" logged into the system"
        elif event.event_type == 'logout':
            description += f" logged out of the system"
        elif event.event_type == 'quiz_action':
            if event.additional_data:
                try:
                    data = json.loads(event.additional_data)
                    if 'score' in data:
                        description += f" completed a quiz with score {data['score']}%"
                    elif 'answer_selected' in data:
                        description += f" answered a quiz question"
                except:
                    description += f" performed a quiz action"
            else:
                description += f" performed a quiz action"
        elif event.event_type == 'video_action':
            if event.video_action:
                description += f" performed video action: {event.video_action}"
            else:
                description += f" interacted with video content"
        
        # Add additional context
        if event.additional_data:
            try:
                data = json.loads(event.additional_data)
                if 'button_text' in data:
                    description += f" (Button: '{data['button_text']}')"
                elif 'link_text' in data:
                    description += f" (Link: '{data['link_text']}')"
            except:
                pass
        
        # Origin (web, mobile, etc.)
        origin = 'web'  # Default to web
        
        processed_events.append({
            'timestamp': event.timestamp,
            'event_context': event_context,
            'component': component,
            'event_name': event_name,
            'description': description,
            'origin': origin,
            'ip_address': event.ip_address or 'N/A',
            'username': username or 'Anonymous',
            'email': email or 'N/A'
        })
    
    return render_template('admin_user_activity.html', events=processed_events)

def get_client_ip():
    """Get the client's IP address"""
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    elif request.headers.get('X-Real-IP'):
        return request.headers.get('X-Real-IP')
    else:
        return request.remote_addr

def track_event(event_type, element_id, element_type, user_id=None, additional_data=None):
    """Helper function to track events"""
    event = ClickstreamEvent(
        user_id=user_id or (current_user.id if current_user.is_authenticated else None),
        session_id=session.get('session_id', 'anonymous'),
        event_type=event_type,
        element_id=element_id,
        element_type=element_type,
        page_url=request.url,
        additional_data=json.dumps(additional_data) if additional_data else None,
        ip_address=get_client_ip()
    )
    db.session.add(event)
    db.session.commit()

# Initialize database and create sample data
def create_tables():
    try:
        with app.app_context():
            db.create_all()
            print("Database tables created successfully!")
            
            # Create sample data if database is empty
            if not Course.query.first():
                # Create sample course
                course = Course(title='Introduction to Python Programming', 
                               description='Learn the basics of Python programming language')
                db.session.add(course)
                db.session.commit()
                
                # Create sample lessons
                lesson1 = Lesson(title='What is Python?', content_type='text', 
                                content='Python is a high-level programming language that emphasizes code readability and simplicity. It was created by Guido van Rossum and first released in 1991. Python supports multiple programming paradigms including procedural, object-oriented, and functional programming.', 
                                course_id=course.id, order=1)
                lesson2 = Lesson(title='Python Basics', content_type='video', 
                                content='https://sample-videos.com/zip/10/mp4/SampleVideo_1280x720_1mb.mp4', 
                                course_id=course.id, order=2)
                lesson3 = Lesson(title='Python Quiz', content_type='quiz', 
                                content='Test your knowledge about Python programming', 
                                course_id=course.id, order=3)
                
                db.session.add_all([lesson1, lesson2, lesson3])
                db.session.commit()
                
                # Create quiz questions
                q1 = QuizQuestion(question='What is Python?', 
                                 options=json.dumps(['A snake', 'A programming language', 'A game', 'A book']),
                                 correct_answer=1, lesson_id=lesson3.id)
                q2 = QuizQuestion(question='Python was created by?', 
                                 options=json.dumps(['Bill Gates', 'Guido van Rossum', 'Steve Jobs', 'Mark Zuckerberg']),
                                 correct_answer=1, lesson_id=lesson3.id)
                
                db.session.add_all([q1, q2])
                db.session.commit()
                
                # Create admin user
                admin_password = generate_password_hash('admin123')
                admin_user = AdminUser(
                    username='admin',
                    email='admin@learningwebsite.com',
                    password_hash=admin_password,
                    is_admin=True
                )
                db.session.add(admin_user)
                db.session.commit()
                print("Sample data and admin user created successfully!")
    except Exception as e:
        print(f"Error creating database: {e}")
        import traceback
        traceback.print_exc()

# Initialize database when app starts
if __name__ == '__main__':
    create_tables()
    app.run(debug=True)
